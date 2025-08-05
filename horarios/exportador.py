import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from horarios.models import Curso, Horario, BloqueHorario

DIAS = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes']

def exportar_horarios_excel(ruta='horarios_generados.xlsx'):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_descanso = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # amarillo
    fill_almuerzo = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # rojo claro

    bloques = list(BloqueHorario.objects.all().order_by('numero'))
    cursos = Curso.objects.all()

    for curso in cursos:
        ws = wb.create_sheet(title=curso.nombre)

        # Encabezado horizontal: días
        for col, dia in enumerate(DIAS, start=2):
            cell = ws.cell(row=1, column=col, value=dia.capitalize())
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Encabezado vertical: bloques
        for row, bloque in enumerate(bloques, start=2):
            cell = ws.cell(row=row, column=1, value=f"Bloque {bloque}")
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Llenar celdas con el horario del curso
        horarios = Horario.objects.filter(curso=curso)
        for bloque in bloques:
            row = bloques.index(bloque) + 2  # Fila para el bloque

            if bloque.tipo != 'clase':
                fill = fill_descanso if bloque.tipo == 'descanso' else fill_almuerzo
                texto = 'Descanso' if bloque.tipo == 'descanso' else 'Almuerzo'
                for col in range(2, len(DIAS) + 2):
                    cell = ws.cell(row=row, column=col, value=texto)
                    cell.alignment = center
                    cell.border = border
                    cell.fill = fill
            else:
                for dia in DIAS:
                    col = DIAS.index(dia) + 2
                    h = horarios.filter(dia=dia, bloque=bloque.numero).first()

                    if h:
                        valor = f"{h.materia.nombre}\n{h.profesor.nombre}\n{h.aula.nombre}"
                        cell = ws.cell(row=row, column=col, value=valor)
                        cell.alignment = center
                        cell.border = border

            # Ajustar tamaña columna
        for col in range(1, len(DIAS) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
        for row in range(1, len(bloques) + 2):
            ws.row_dimensions[row].height = 60

    wb.save(ruta)
    print(f"✅ Archivo Excel generado: {ruta}")


